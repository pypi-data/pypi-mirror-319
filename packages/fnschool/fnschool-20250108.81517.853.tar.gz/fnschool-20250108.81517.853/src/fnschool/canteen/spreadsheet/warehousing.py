import os
import sys
from openpyxl.styles import *
from openpyxl.formatting.rule import *
from openpyxl.styles.differential import *
from openpyxl.utils.cell import *


from fnschool import *

from fnschool.canteen.spreadsheet.base import *


class Warehousing(Base):
    def __init__(self, bill, form_row_len=21):
        super().__init__(bill)
        self.sheet_name = self.s.warehousing_name
        self.entry_row_len0 = 21
        pass

    def get_entry_index(self, form_index):
        form_index0, form_index1 = form_index
        entry_index = [form_index0 + 2, form_index1 - 1]
        return entry_index

    def format(self):
        wsheet = self.sheet
        self.unmerge_sheet_cells()

        for row in wsheet.iter_rows(
            min_row=1, max_row=wsheet.max_row, min_col=1, max_col=8
        ):
            if row[0].value and row[0].value.replace(" ", "") == "入库单":
                wsheet.row_dimensions[row[0].row].height = 21
                wsheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )
                wsheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=4,
                    end_column=6,
                )
                wsheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=7,
                    end_column=8,
                )

            if row[0].value and row[0].value.replace(" ", "").endswith("类"):
                row[6].number_format = numbers.FORMAT_NUMBER_00
                for _row in wsheet.iter_rows(
                    min_row=row[0].row + 1,
                    max_row=wsheet.max_row + 1,
                    min_col=1,
                    max_col=1,
                ):
                    wsheet.row_dimensions[row[0].row].height = self.row_height
                    if _row[0].value and (
                        _row[0].value.replace(" ", "").endswith("类")
                        or _row[0].value.replace(" ", "") == "合计"
                    ):
                        wsheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=1,
                            end_column=1,
                        )
                        wsheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=7,
                            end_column=7,
                        )
                        break

                ci0, ci1 = row[0].row, 0
                for _row in wsheet.iter_rows(
                    min_row=row[0].row + 1,
                    max_row=wsheet.max_row + 1,
                    min_col=1,
                    max_col=7,
                ):
                    if _row[0].value and (
                        str(_row[0].value).endswith("类")
                        or "合计" in str(_row[0].value)
                    ):

                        ci1 = _row[0].row - 1

                        fnames = [
                            wsheet.cell(r, 2).value for r in range(ci0, ci1 + 1)
                        ]
                        fnames_d = list(
                            set(
                                [n for n in fnames if fnames.count(n) > 1 and n]
                            )
                        )

                        for dn in fnames_d:
                            row_index0_d = row[0].row + fnames.index(dn)
                            row_index1_d = row_index0_d + fnames.count(dn) - 1
                            fname = dn
                            funit_name = wsheet.cell(row_index0_d, 3).value
                            ftotal = sum(
                                [
                                    wsheet.cell(r, 6).value
                                    for r in range(
                                        row_index0_d, row_index1_d + 1
                                    )
                                ]
                            )

                            for r in range(row_index0_d, row_index1_d + 1):
                                wsheet.cell(r, 2, "")
                                wsheet.cell(r, 3, "")
                                wsheet.cell(r, 6, "")

                            wsheet.merge_cells(
                                start_row=row_index0_d,
                                end_row=row_index1_d,
                                start_column=2,
                                end_column=2,
                            )
                            wsheet.merge_cells(
                                start_row=row_index0_d,
                                end_row=row_index1_d,
                                start_column=3,
                                end_column=3,
                            )
                            wsheet.merge_cells(
                                start_row=row_index0_d,
                                end_row=row_index1_d,
                                start_column=6,
                                end_column=6,
                            )
                            wsheet.cell(row_index0_d, 2, fname)
                            wsheet.cell(row_index0_d, 3, funit_name)
                            wsheet.cell(row_index0_d, 6, ftotal)

                        break

            if row[0].value and "审核人" in row[0].value.replace(" ", ""):
                wsheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )

        wb = self.bwb
        wb.active = wsheet

        print_info(_("Sheet '%s' was formatted.") % self.sheet.title)

    @property
    def form_indexes(self):
        if self._form_indexes:
            return self._form_indexes
        wsheet = self.sheet
        indexes = []
        row_index = 1
        for row in wsheet.iter_rows(max_row=wsheet.max_row + 1, max_col=8):
            if row[0].value:
                if row[0].value.replace(" ", "") == "入库单":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            self._form_indexes = indexes
            return self._form_indexes

        return None

    def update(self):
        wsheet = self.sheet
        foods = [
            f
            for f in self.bfoods
            if (not f.is_inventory and not f.is_abandoned)
        ]
        form_indexes = self.form_indexes
        class_names = self.bill.food_class_names

        self.unmerge_sheet_cells(wsheet)

        for form_index0, form_index1 in form_indexes:
            food_index0 = form_index0 + 2
            food_index1 = form_index1 - 1
            for row in wsheet.iter_rows(
                min_row=food_index0,
                max_row=food_index1,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""

        w_times = sorted(list(set([f.xdate for f in foods])))

        max_time_index = 0
        for windex, w_time in enumerate(w_times):
            max_time_index = windex + 1
            form_index0, form_index1 = form_indexes[windex]
            food_index0 = form_index0 + 2
            food_index1 = form_index1 - 1
            entry_index = food_index0
            warehousing_n = windex + 1

            wfoods = [f for f in foods if (f.xdate == w_time)]
            w_class_names = [f.fclass for f in wfoods]
            w_class_names_no_food = [
                name for name in class_names if not name in w_class_names
            ]
            row_difference = (
                len(wfoods)
                + len(w_class_names_no_food)
                - (food_index1 - food_index0 + 1)
            )

            if row_difference > 0:
                self.row_inserting_tip(food_index0 + 1)
                wsheet.insert_rows(food_index0 + 1, row_difference)
                for row in wsheet.iter_rows(
                    min_row=food_index0 + 1,
                    max_row=food_index0 + 1 + row_difference,
                    min_col=1,
                    max_col=8,
                ):
                    for cell in row:
                        cell.alignment = self.cell_alignment0
                        self.border = self.cell_border0

                self.del_form_indexes()
                form_indexes = self.form_indexes

                form_index0, form_index1 = form_indexes[windex]
                food_index0 = form_index0 + 2
                food_index1 = form_index1 - 1
                entry_index = food_index0

                row_difference = 0

            for row in wsheet.iter_rows(
                min_row=food_index0,
                max_row=food_index1,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

            wsheet.cell(form_index0, 2, self.purchaser)
            wsheet.cell(
                form_index0,
                4,
                f"{w_time.year}年 {w_time.month} 月 "
                + f"{w_time.day} 日  单位：元",
            )
            wsheet.cell(
                form_index0,
                7,
                f"编号：R{w_time.month:0>2}{warehousing_n:0>2}",
            )

            for class_name in class_names:
                cfoods = [f for f in wfoods if f.fclass == class_name]
                cfoods = sorted(cfoods, key=lambda f: f.name)

                cfoods_total_price = sum([f.total_price for f in cfoods])

                wsheet.cell(entry_index, 1, class_name)
                wsheet.cell(entry_index, 7, cfoods_total_price)

                if len(cfoods) < 1:
                    entry_index += 1
                    continue

                for cfindex, cfood in enumerate(cfoods):
                    cfood_row_index = entry_index + cfindex
                    wsheet.cell(cfood_row_index, 2, cfood.name)
                    wsheet.cell(cfood_row_index, 3, cfood.unit_name)
                    wsheet.cell(cfood_row_index, 4).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )
                    wsheet.cell(cfood_row_index, 4, cfood.count)
                    wsheet.cell(cfood_row_index, 5).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )
                    wsheet.cell(cfood_row_index, 5, cfood.unit_price)
                    wsheet.cell(cfood_row_index, 6).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )
                    wsheet.cell(cfood_row_index, 6, cfood.total_price)

                entry_index_end = entry_index + len(cfoods) - 1

                if class_name == class_names[0] and row_difference < 0:
                    entry_index_end = (
                        entry_index_end
                        + abs(row_difference)
                        - len(w_class_names_no_food)
                    )

                entry_index = entry_index_end + 1
            foods_total_price = sum([f.total_price for f in wfoods])
            wsheet.cell(form_index1, 6, foods_total_price)
            wsheet.cell(form_index1, 7, foods_total_price)

            wsheet.cell(
                form_index1 + 1,
                1,
                (
                    "   "
                    + "审核人："
                    + "        "
                    + "经办人："
                    + self.operator.name
                    + " 　    "
                    + "过称人："
                    + "        "
                    + "仓管人："
                    + " 　"
                ),
            )

        if len(form_indexes) > max_time_index:
            for time_index in range(max_time_index, len(form_indexes)):
                form_index0, form_index1 = form_indexes[time_index]
                food_index0, food_index1 = (
                    form_index0 + 2,
                    form_index1 - 1,
                )

                for row in wsheet.iter_rows(
                    min_row=food_index0,
                    max_row=food_index1,
                    min_col=2,
                    max_col=7,
                ):
                    for cell in row:
                        cell.value = ""

        self.del_form_empty_rows([1, 2])
        self.format()
        wb = self.bwb
        wb.active = wsheet

        print_info(_("Sheet '%s' was updated.") % (self.sheet.title))


# The end.
