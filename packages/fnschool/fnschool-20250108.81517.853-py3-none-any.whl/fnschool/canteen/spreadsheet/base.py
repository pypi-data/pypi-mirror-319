import os
import sys
from openpyxl.styles import *
from openpyxl.formatting.rule import *
from openpyxl.styles.differential import *
from openpyxl.utils.cell import *


from fnschool import *


class Base:
    def __init__(self, bill):
        self.bill = bill
        self.currency = self.bill.currency
        self.spreadsheet = self.bill.spreadsheet
        self.s = self.spreadsheet
        self._bill_workbook = None
        self.sd = self.bill.significant_digits
        self._bwb = None
        self.sheet_name = None
        self._sheet = None
        self._form_indexes = None
        self.cell_alignment0 = Alignment(horizontal="center", vertical="center")
        self.cell_side0 = Side(border_style="thin")
        self.cell_border0 = Border(
            top=self.cell_side0,
            left=self.cell_side0,
            right=self.cell_side0,
            bottom=self.cell_side0,
        )
        self.cell_side1 = Side(border_style="thin", color="ff0066")
        self.cell_border1 = Border(
            top=self.cell_side1,
            left=self.cell_side1,
            right=self.cell_side1,
            bottom=self.cell_side1,
        )
        self.cell_fill0 = PatternFill("solid", start_color="ffccff")
        self.row_height = 12.75
        self.filetypes_xlsx = [(_("Spreadsheet Files"), "*.xlsx")]
        self.food_sheet0_name = "材料台账母表"
        self.food_form_title_like = "材料入库、出库台账"

    def get_local_total_price(self, total_price):
        return (
            self.bill.get_CNY_chars(total_price)
            if is_zh_CN
            else self.bill.get_CNY_chars(total_price)
        )

    @property
    def bill_workbook(self):
        if not self._bill_workbook:
            self._bill_workbook = self.spreadsheet.bill_workbook
        return self._bill_workbook

    @property
    def bwb(self):
        if not self._bwb:
            self._bwb = self.bill_workbook
        return self._bwb

    def row_inserting_tip(self, row_index):
        print_error(
            _(
                "Row {0} of {1} is being inserted, " + "please wait a moment."
            ).format(row_index, self.sheet.title)
        )

    def del_form_indexes(self):
        self._form_indexes = None

    @property
    def consuming_day_m1(self):
        dates = []
        for f in self.bfoods:
            dates += [d for d, __ in f.consumptions]
        date = max(dates)
        return date.day

    @property
    def bill_foods(self):
        return self.bill.foods

    @property
    def bfoods(self):
        return self.bill_foods

    @property
    def purchaser(self):
        return self.bill.purchaser

    @property
    def operator(self):
        return self.bill.operator

    @property
    def config(self):
        return self.operator.config

    def get_bill_sheet(self, name):
        sheet = self.bwb[name]
        return sheet

    def unmerge_sheet_cells(self, sheet=None):
        sheet = sheet or self.sheet
        if isinstance(sheet, str):
            sheet = self.get_bill_sheet(sheet)
        merged_ranges = list(sheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            sheet.unmerge_cells(str(cell_group))
        print_info(_("Cells of {0} was unmerged.").format(sheet.title))

    @property
    def sheet(self):
        if not self.sheet_name:
            return None
        if not self._sheet:
            self._sheet = self.get_bill_sheet(self.sheet_name)
        return self._sheet

    def get_bill_year(self, wb=None):
        return (
            self.get_zh_CN_bill_year(wb)
            if is_zh_CN
            else self.get_zh_CN_bill_year(wb)
        )

    def get_zh_CN_bill_year(self, wb=None):
        return self.get_bill_year_by_x_year(wb, "年")

    def get_bill_year_by_x_year(self, wb=None, year_word="年"):
        wb = wb or self.bwb
        cover_sheet = wb[self.s.cover_name]
        year = str(cover_sheet.cell(1, 1).value)
        year0 = ""
        year_word = year_word
        year = year.split(year_word)[0]
        for i in range(len(year) - 1, -1, -1):
            if year[i].isnumeric():
                year0 = year[i] + year0
            else:
                break

        bill_year = year0

        return bill_year

    def update_food_sheet_year(self, sheet):
        wb = sheet.parent
        bill_year = self.get_bill_year(wb)
        csheet = sheet
        for row_index in range(1, csheet.max_row + 1):
            cell0_value = str(csheet.cell(row_index, 1).value).replace(" ", "")
            if cell0_value.endswith("年") and cell0_value[:-1].isnumeric():
                csheet.cell(row_index, 1, bill_year + "年")
            pass
        pass

    def del_form_empty_rows(self, empty_cols):

        self.del_form_indexes()
        form_indexes = self.form_indexes

        empty_cols = (
            [empty_cols] if not isinstance(empty_cols, list) else empty_cols
        )

        form_indexes_len = len(form_indexes)
        for i in range(form_indexes_len):

            self.del_form_indexes()

            form_index = self.form_indexes[i]
            entry_index0, entry_index1 = self.get_entry_index(form_index)

            entry_len = (entry_index1 - entry_index0) + 1
            len_diff = entry_len - self.entry_row_len0

            if len_diff > 0:

                entry_index1_0 = entry_index1 + 1
                for row_index in range(entry_index0, entry_index1 + 1):
                    if all(
                        [
                            self.sheet.cell(row_index, col_index).value is None
                            for col_index in empty_cols
                        ]
                    ):
                        self.sheet.delete_rows(row_index, 1)
                        print_warning(
                            _(
                                'Empty row {0} of sheet "{1}" has been deleted.'
                            ).format(row_index, self.sheet.title)
                        )
                        entry_index1_0 -= 1
                        len_diff -= 1

                        if row_index >= entry_index1_0 or len_diff < 1:
                            break
                            pass
                        pass
                    pass
                pass
            pass
        pass

    pass


# The end.
