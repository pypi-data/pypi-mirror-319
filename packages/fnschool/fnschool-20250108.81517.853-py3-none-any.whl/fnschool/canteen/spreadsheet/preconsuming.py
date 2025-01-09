import os
import sys
import random
import calendar
from datetime import datetime, timedelta
import math
from openpyxl.styles import PatternFill

from fnschool import *
from fnschool.canteen.food import *
from fnschool.canteen.path import *
from fnschool.canteen.spreadsheet.base import *


class PreConsuming(Base):
    def __init__(self, bill):
        super().__init__(bill)
        self.path0 = pre_consuming0_fpath
        self.row_index_offset = 3
        self.col_index_offset = 5

        self.sheet_name0 = self.s.preconsuming_name0

    def get_colored_cols(self, days_diff):
        days_diff += 4
        colored_cols = list(range(1, days_diff + 1))
        colored_cols.remove(3)
        return colored_cols

    def pre_consume_foods(self):
        foods = self.bill.foods
        cfoods = [f for f in foods if not f.is_abandoned]
        if len(cfoods) < 1:
            print_error(_("No food found, exit."))
            exit()
        year = self.bill.consuming.year
        month = self.bill.consuming.month
        __, last_day = calendar.monthrange(year, month)
        mday_m1 = datetime(year, month, last_day)
        time_nodes = sorted(
            list(
                set(
                    [f.xdate for f in cfoods]
                    + [
                        datetime(
                            year,
                            month,
                            calendar.monthrange(year, month)[1],
                        )
                    ]
                )
            )
        )
        if len(time_nodes) > 1:
            t1 = time_nodes[1]
            if t1.day == 1:
                del time_nodes[1]

        wb_fpathes = []
        for i in range(1, len(time_nodes)):
            tn0, tn1 = time_nodes[i - 1], time_nodes[i]
            tn0_cp = tn0
            if tn0.month != tn1.month:
                tn0 = datetime(tn1.year, tn1.month, 1)

            file_t0 = (
                tn0
                if not tn0_cp.month == tn1.month
                else (tn0 + timedelta(days=1))
            )
            file_t1 = tn1

            file_t0 = _("{year}.{month}.{day}").format(
                year=file_t0.year, month=file_t0.month, day=file_t0.day
            )
            file_t1 = _("{year}.{month}.{day}").format(
                year=file_t1.year, month=file_t1.month, day=file_t1.day
            )

            wb_fpath = (
                self.bill.operator.preconsuming_dpath
                / (
                    _("food_consuming--{t0}--{t1}.xlsx").format(
                        t0=file_t0, t1=file_t1
                    )
                )
            ).as_posix()

            wb_fpathes.append(wb_fpath)

        for i, wb_fpath in enumerate(wb_fpathes):
            if not Path(wb_fpath).exists():
                shutil.copy(self.path0, wb_fpath)
                print_info(
                    _('Spreadsheet "{0}" was copied to "{1}".').format(
                        self.path0, wb_fpath
                    )
                )
            wb = load_workbook(wb_fpath)
            sheet = wb[self.sheet_name0]
            tn1 = time_nodes[i + 1]
            tn0 = time_nodes[i]
            tn0_cp = tn0

            new_meal_type_rows = []

            if not tn0.month == tn1.month:
                tn0 = datetime(tn1.year, tn1.month, 1)

            wbfoods = [
                f
                for f in cfoods
                if f.get_remainder(mday_m1) > 0 and f.xdate <= tn0
            ]

            wbfoods = sorted(
                wbfoods,
                key=lambda f: (f.meal_type, f.xdate, f.name, f.unit_price),
            )

            col_index = 0
            tn1_r = calendar.monthrange(tn1.year, tn1.month)
            tn1_r = datetime(tn1.year, tn1.month, tn1_r[-1])
            days_diff = (tn1_r - tn0).days

            for d_index in range(0, days_diff + 1):
                d_date = tn0 + timedelta(
                    days=d_index + (1 if tn0_cp.month == tn1_r.month else 0)
                )
                col_index = self.col_index_offset + d_index
                sheet.cell(
                    1,
                    col_index,
                    d_date.strftime("%Y.%m.%d"),
                )
                if d_date == tn1_r:
                    break

            for col_index in range(col_index + 1, sheet.max_column):
                sheet.cell(1, col_index, "")

            row_index = 0
            last_meal_type = None
            for f_index in range(0, len(wbfoods)):
                wbfood = wbfoods[f_index]
                row_index = self.row_index_offset + f_index
                sheet.cell(
                    row_index, 1, wbfood.get_display_name(time_node0=tn0)
                )
                sheet.cell(row_index, 2, wbfood.get_remainder(mday_m1))
                sheet.cell(row_index, 4, wbfood.unit_price)
                if wbfood.meal_type and not last_meal_type == wbfood.meal_type:
                    new_meal_type_rows.append([wbfood.meal_type, row_index])
                    last_meal_type = wbfood.meal_type

            for row_index in range(row_index + 1, sheet.max_row + 1):
                sheet.cell(row_index, 1, "")
                sheet.cell(row_index, 2, "")
                sheet.cell(row_index, 4, "")
                pass

            new_meal_type_rows = [
                [new_meal_type_rows[i][1], new_meal_type_rows[i + 1][1]]
                for i in range(0, len(new_meal_type_rows), 2)
                if i + 1 < len(new_meal_type_rows)
            ]

            for [row_index0, row_index1] in new_meal_type_rows:
                colored_cols = self.get_colored_cols(days_diff)

                for row_index in range(row_index0, row_index1):
                    for col_index in colored_cols:
                        cell = sheet.cell(row_index, col_index)
                        cell.fill = self.cell_fill0
                        cell.border = self.cell_border1

            wb.save(wb_fpath)
            print_warning(
                _(
                    "Sheet '{0}' of \"{1}\" was updated.\n"
                    + "Press any key to continue when you have "
                    + "completed the foods allocation."
                ).format(sheet.title, wb_fpath)
            )
            new_wbfoods = [
                f for f in cfoods if f.get_remainder(tn1) > 0 and f.xdate == tn1
            ]
            if len(new_wbfoods) > 0:
                print_info(
                    (
                        _("New purchased food for date {0} is:")
                        if len(new_wbfoods) > 1
                        else _("New purchased foods for date {0} are:")
                    ).format(tn1.strftime("%Y.%m.%d"))
                )

                new_wbfoods_count_len = len(str(len(new_wbfoods)))

                number_mark = random.choice([" ", ".", ">", "-", ")", ":", "|"])
                new_wbfood_tips = [
                    (
                        f"{i+1:>{new_wbfoods_count_len}}"
                        + number_mark
                        + f"{f.name}"
                        + _("({0})").format(f.meal_type)
                        + f" {f.count} {f.unit_name}"
                    )
                    for i, f in enumerate(new_wbfoods)
                ]

                new_wbfood_tips_value = sqr_slist(new_wbfood_tips)

                print_warning(new_wbfood_tips_value)

                print_warning(_("Negligible foods are not listed."))
            else:
                print_info(
                    _("There is no purchased food for {0}.").format(
                        tn1.strftime("%Y.%m.%d")
                    )
                )

            print_error(
                _(
                    "There is no need to design for "
                    + "dates without food consumption. "
                    + "(Ok, I know [press any key to continue])"
                )
            )
            get_input()
            wb.close()
            open_path(wb_fpath)
            print_info(
                _(
                    "Ok! I have updated spreadsheet '{0}'. (Press any key)"
                ).format(wb_fpath)
            )
            get_input()

            wb = load_workbook(wb_fpath)
            sheet = wb[self.sheet_name0]

            f_index = 0
            for row in sheet.iter_rows(
                min_row=self.row_index_offset,
                min_col=self.col_index_offset,
                max_row=sheet.max_row,
                max_col=sheet.max_column,
            ):
                if f_index == len(wbfoods):
                    break
                food = wbfoods[f_index]
                col_index = self.col_index_offset
                for cell in row:
                    if cell.value:
                        cdate = sheet.cell(1, col_index).value
                        food.consumptions.append(
                            [
                                datetime.strptime(cdate, "%Y.%m.%d"),
                                float(cell.value),
                            ]
                        )
                    col_index += 1
                f_index += 1
            wb.close()
            sheet = None
        self.print_consuming_days(cfoods)
        pass

    def print_consuming_days(self, foods):
        consumption_days = []
        year = self.bill.consuming.year
        month = self.bill.consuming.month
        year_month = _("{year}.{month}").format(year=year, month=month)

        meal_types = list(set([f.meal_type for f in foods]))

        for meal_type in meal_types:

            for f in foods:
                for d, c in f.consumptions:
                    if not d.day in consumption_days:
                        consumption_days.append(d.day)

            consumption_days_value = ""
            space_len = 5
            for week in calendar.monthcalendar(year, month):
                for d in week:
                    if d == 0:
                        consumption_days_value += " " * space_len
                    elif d in consumption_days:
                        consumption_day = f"({d:>2})"
                        consumption_days_value += (
                            f"{consumption_day:>{space_len}}"
                        )
                    else:
                        d = f" {d:>2} "
                        consumption_days_value += f"{d:>{space_len}}"
                consumption_days_value += "\n"

            print()
            print_error(
                _("Consuming days of {0} ({1}):").format(year_month, meal_type)
                if meal_type
                else _("Consuming days of {0}:").format(year_month)
            )
            print_warning(f"{year_month:^{space_len*7}}")
            if consumption_days_value.endswith("\n"):
                consumption_days_value = consumption_days_value[:-1]
            print_info(consumption_days_value[:-1])
            consumption_days_len = len(consumption_days)
            total_days = (
                _("{0} days in total.")
                if consumption_days_len > 1
                else _("{0} day in total.")
            ).format(consumption_days_len)
            print_warning(f"{total_days:^{space_len*7}}")
            print_info(
                _("Yes, they are all right. (Press any key to continue)")
            )

            get_input()

            pass

        return


# The end.
