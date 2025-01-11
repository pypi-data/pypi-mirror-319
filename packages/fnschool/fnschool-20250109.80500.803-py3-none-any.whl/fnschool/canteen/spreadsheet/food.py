import os
import sys
import secrets
import random

from fnschool import *
from fnschool.canteen.spreadsheet.base import *
from openpyxl.styles import *


class Food(Base):
    def __init__(self, bill):
        super().__init__(bill)
        self.sheet_name = self.s.sfood_name
        self._food_sheet0 = None
        pass

    def get_sheet(self, name=None, wb=None):
        sheet = None
        bfoods = self.bfoods
        wb = wb or self.bwb

        if name in wb.sheetnames:
            sheet = wb[name]
        else:
            sheet = wb.copy_worksheet(wb[self.sheet_name])
            sheet.title = name

        for row_index in range(1, sheet.max_row + 1):
            rc1_value = sheet.cell(row_index, 1).value
            rc1_value = str(rc1_value)
            if rc1_value and "材料名称：（）" in rc1_value:
                unit = [f.unit_name for f in bfoods if f.name == name]
                unit = unit[0] if len(unit) > 0 else "斤"
                sheet.cell(row_index, 1, f"材料名称：{name}（{unit}）")

        return sheet

    def format(self, sheet):
        if isinstance(sheet, str):
            sheet = self.get_sheet(sheet)

        self.unmerge_sheet_cells(sheet)

        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=14,
        ):
            sheet.row_dimensions[row[0].row].height = 15.75
            if row[0].value and "入库、出库台账" in str(row[0].value):
                sheet.row_dimensions[row[0].row].height = 27
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=13,
                )
                row[0].border = self.cell_border0

            if row[0].value and "年" in str(row[0].value):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=2,
                )
                row[0].border = self.cell_border0

            if row[3].value and "入库" in str(row[3].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=4,
                    end_column=6,
                )
                row[3].border = self.cell_border0

            if row[6].value and "出库" in str(row[6].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=7,
                    end_column=9,
                )
                row[6].border = self.cell_border0

            if row[9].value and "库存" in str(row[9].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=10,
                    end_column=12,
                )
                row[9].border = self.cell_border0

            if row[12].value and "编号" in str(row[12].value).replace(" ", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=13,
                    end_column=13,
                )
                row[12].border = self.cell_border0

            if row[0].value and self.food_form_title_like in str(row[0].value):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=13,
                )
                row[0].font = Font(size=18, bold=True)
                row[0].alignment = self.cell_alignment0
                row[0].border = self.cell_border0
                if row[0].row - 1 > 0:
                    note_cell = sheet.cell(row[0].row - 1, 2)
                    note_cell.value = (
                        "注：《学校食堂材料入库、出库台账》"
                        + "是以入库单、出库单为依据按日进行登记。"
                    )
                    note_cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )
                    note_cell.border = self.cell_border0
                    sheet.merge_cells(
                        start_row=note_cell.row,
                        end_row=note_cell.row,
                        start_column=2,
                        end_column=13,
                    )

        print_info(_("Sheet {0} has been reformatted.").format(sheet.title))

    def get_form_index(self, sheet):
        indexes = self.get_form_indexes(sheet)
        index_range = indexes[self.bill.consuming.month - 1]
        return index_range

    def get_form_indexes(self, sheet):
        indexes = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
        ):
            if row[0].value and "材料名称" in str(row[0].value).replace(
                " ", ""
            ):
                indexes.append([row[0].row + 3, None])

            if row[2].value and "本月合计" in str(row[2].value).replace(
                " ", ""
            ):
                indexes[-1][1] = row[0].row + 1
        return indexes

    def update_inventories(self, sheet):

        print_warning(_('Updating inventories for "{0}"').format(sheet.title))

        row_indexes = []
        row_index0 = None
        row_index1 = None
        for row in sheet.iter_rows(max_col=14):
            if str(row[2].value) == "摘要":
                row_index0 = row[2].row + 1

            elif str(row[2].value) == "本月合计":
                row_index1 = row[2].row - 1
                if not row_index0:
                    print_error(
                        _(
                            "cannot find the starting index of form in "
                            + 'Sheet "{0}". (row {1})'
                        ).format(sheet.title, row[0].row)
                    )
                if not row_index1:
                    print_error(
                        _(
                            "cannot find the ending index of form in "
                            + 'Sheet "{0}". (row {1})'
                        ).form(sheet.title, row[0].row)
                    )
                if row_index0 and row_index1:
                    row_indexes.append([row_index0, row_index1])
                row_index0 = None
                row_index1 = None
            pass

        for ri0, ri1 in row_indexes:

            unit_prices = [
                sheet.cell(ri, 11).value for ri in range(ri0, ri1 + 1)
            ]
            unit_prices = [
                float(u)
                for u in unit_prices
                if str(u).replace(".", "").isnumeric()
            ]

            unit_price_rows = []
            for u in unit_prices:
                u_rows = []
                for ri in range(ri0, ri1 + 1):
                    cell10_value = sheet.cell(ri, 11).value
                    if str(cell10_value).replace(".", "").isnumeric():
                        if float(cell10_value) == u:
                            u_rows.append(ri)
                unit_price_rows.append([u, u_rows])

            for unit_price, rows in unit_price_rows:
                rows_len = len(rows)

                for row_i in range(rows_len):
                    if row_i + 1 >= rows_len:
                        break
                    row_index = rows[row_i]
                    row_indexp1 = rows[row_i + 1]

                    w_count_n = sheet.cell(row_indexp1, 4).value
                    w_unit_price_n = sheet.cell(row_indexp1, 5).value
                    if str(w_count_n).replace(".", "").isnumeric():
                        w_count_n = float(w_count_n)
                    else:
                        w_count_n = 0
                    if str(w_unit_price_n).replace(".", "").isnumeric():
                        w_unit_price_n = float(w_unit_price_n)
                    else:
                        w_unit_price_n = 0
                    w_total_price_n = w_count_n * w_unit_price_n

                    c_count_n = sheet.cell(row_indexp1, 7).value
                    c_unit_price_n = sheet.cell(row_indexp1, 8).value
                    if str(c_count_n).replace(".", "").isnumeric():
                        c_count_n = float(c_count_n)
                    else:
                        c_count_n = 0
                    if str(c_unit_price_n).replace(".", "").isnumeric():
                        c_unit_price_n = float(c_unit_price_n)
                    else:
                        c_unit_price_n = 0
                    c_total_price_n = c_count_n * c_unit_price_n

                    i_count_n = sheet.cell(row_indexp1, 10).value
                    i_unit_price_n = sheet.cell(row_indexp1, 11).value
                    if str(i_count_n).replace(".", "").isnumeric():
                        i_count_n = float(i_count_n)
                    else:
                        i_count_n = 0
                    if str(i_unit_price_n).replace(".", "").isnumeric():
                        i_unit_price_n = float(i_unit_price_n)
                    else:
                        i_unit_price_n = 0
                    i_total_price_n = i_count_n * i_unit_price_n

                    i_count = sheet.cell(row_index, 10).value
                    i_unit_price = sheet.cell(row_index, 11).value
                    if str(i_count).replace(".", "").isnumeric():
                        i_count = float(i_count)
                    else:
                        i_count = 0
                    if str(i_unit_price).replace(".", "").isnumeric():
                        i_unit_price = float(i_unit_price)
                    else:
                        i_unit_price = 0
                    i_total_price = i_count * i_unit_price

                    if (
                        i_unit_price == w_unit_price_n
                        and not i_count_n == (i_count + w_count_n - c_count_n)
                        and not (w_count_n == 0 and c_count_n == 0)
                    ):
                        sheet.cell(
                            row_indexp1, 10, i_count + w_count_n - c_count_n
                        )
                        sheet.cell(
                            row_indexp1,
                            12,
                            i_unit_price * (i_count + w_count_n - c_count_n)
                        )

                        pass
                    pass
                pass

            for row_index in range(ri0,ri1+1):
                w_count = sheet.cell(row_index, 4).value
                w_unit_price = sheet.cell(row_index, 5).value
                if str(w_count).replace(".", "").isnumeric():
                    w_count = float(w_count)
                else:
                    w_count = 0
                if str(w_unit_price).replace(".", "").isnumeric():
                    w_unit_price = float(w_unit_price)
                else:
                    w_unit_price = 0
                w_total_price = w_count * w_unit_price
                
                if not sheet.cell(row_index,6).value:
                    sheet.cell(row_index, 6, w_total_price)

                c_count = sheet.cell(row_index, 7).value
                c_unit_price = sheet.cell(row_index, 8).value
                if str(c_count).replace(".", "").isnumeric():
                    c_count = float(c_count)
                else:
                    c_count = 0
                if str(c_unit_price).replace(".", "").isnumeric():
                    c_unit_price = float(c_unit_price)
                else:
                    c_unit_price = 0
                c_total_price = c_count * c_unit_price

                if not sheet.cell(row_index, 9).value:
                    sheet.cell(row_index, 9, c_total_price)

                i_count = sheet.cell(row_index, 10).value
                i_unit_price = sheet.cell(row_index, 11).value
                if str(i_count).replace(".", "").isnumeric():
                    i_count = float(i_count)
                else:
                    i_count = 0
                if str(i_unit_price).replace(".", "").isnumeric():
                    i_unit_price = float(i_unit_price)
                else:
                    i_unit_price = 0
                i_total_price = i_count * i_unit_price

                if i_count == 0:
                    sheet.cell(row_index, 10, "0")

                if not sheet.cell(row_index, 12).value:
                    sheet.cell(row_index, 12, i_total_price)




            pass

        print_info(_("Update completed!"))

        pass

    def update_summation(self, sheet):

        print_info(
            _('Update monthly and annual accumulation for "{0}".').format(
                sheet.title
            )
        )
        warehousing_sum_m = 0
        warehousing_sum_y = 0

        warehousing_count_m = 0
        warehousing_count_y = 0

        consuming_sum_m = 0
        consuming_sum_y = 0

        consuming_count_m = 0
        consuming_count_y = 0

        for row in sheet.iter_rows(max_col=14):
            if str(row[3].value) == "数量":
                warehousing_count_m = 0
                warehousing_sum_m = 0

                consuming_count_m = 0
                consuming_sum_m = 0

                current_row_index = row[0].row
                row_index_end = None
                for row_index in range(current_row_index, sheet.max_row + 1):
                    if str(sheet.cell(row_index, 3).value) == "本月合计":
                        row_index_end = row_index
                        break

                for row_index in range(current_row_index, row_index_end):
                    cell3_value = str(sheet.cell(row_index, 4).value).replace(
                        ".", ""
                    )

                    if cell3_value.isnumeric():
                        cell3_value = str(sheet.cell(row_index, 4).value)

                        warehousing_count_m += float(cell3_value)

                        cell4_value = str(
                            sheet.cell(row_index, 5).value
                        ).replace(".", "")

                        if cell4_value.isnumeric():
                            cell4_value = str(sheet.cell(row_index, 5).value)
                            sheet.cell(
                                row_index,
                                6,
                                float(cell3_value) * float(cell4_value),
                            )

                    cell5_value = str(sheet.cell(row_index, 6).value).replace(
                        ".", ""
                    )
                    if cell5_value.isnumeric():
                        cell5_value = str(sheet.cell(row_index, 6).value)

                        warehousing_sum_m += float(cell5_value)

                    cell6_value = str(sheet.cell(row_index, 7).value).replace(
                        ".", ""
                    )
                    if cell6_value.isnumeric():
                        cell6_value = str(sheet.cell(row_index, 7).value)
                        consuming_count_m += float(cell6_value)

                        cell7_value = str(
                            sheet.cell(row_index, 8).value
                        ).replace(".", "")
                        if cell7_value.isnumeric():
                            cell7_value = str(sheet.cell(row_index, 8).value)
                            sheet.cell(
                                row_index,
                                9,
                                float(cell6_value) * float(cell7_value),
                            )

                    cell8_value = str(sheet.cell(row_index, 9).value).replace(
                        ".", ""
                    )
                    if cell8_value.isnumeric():
                        cell8_value = str(sheet.cell(row_index, 9).value)

                        consuming_sum_m += float(cell8_value)

                warehousing_count_y += warehousing_count_m
                warehousing_sum_y += warehousing_sum_m

                consuming_count_y += consuming_count_m
                consuming_sum_y += consuming_sum_m

            if str(row[2].value) == "本年累计":
                row[3].value = warehousing_count_y
                row[5].value = warehousing_sum_y
                row[6].value = consuming_count_y
                row[8].value = consuming_sum_y
                pass

            elif str(row[2].value) == "本月合计":
                row[3].value = warehousing_count_m
                row[5].value = warehousing_sum_m
                row[6].value = consuming_count_m
                row[8].value = consuming_sum_m

                warehousing_count_m = None
                warehousing_sum_m = None
                consuming_sum_m = None
                consuming_count_m = None
                pass

            else:
                pass

        warehousing_sum_y = None
        warehousing_count_y = None
        consuming_sum_y = None
        consuming_count_y = None

        print_warning(_("Update completed."))

    def update(self):

        year = self.bill.consuming.year
        month = self.bill.consuming.month
        cfoods = [f for f in self.bfoods if not f.is_abandoned]
        food_names = list(set([f.name for f in cfoods]))
        wb = self.bwb

        wb[self.sheet_name].sheet_state = "visible"

        rfoods = [
            f for f in self.bfoods if (not f.is_abandoned and f.is_inventory)
        ]

        consuming_days = []
        for f in cfoods:
            for d, c in f.consumptions:
                if (not d in consuming_days) and (d.month == month):
                    consuming_days.append(d)

        consuming_days = sorted(consuming_days)

        warehousing_days = sorted(
            list(set([f.xdate for f in cfoods if f.xdate.month == month]))
        )

        food_names = list(set([f.name for f in rfoods] + food_names))

        sheet = None
        for food_name in food_names:
            sheet = self.get_sheet(food_name)
            form_index_range = self.get_form_index(sheet)
            index_start, index_end = form_index_range

            for row_index in range(index_start, index_end - 1):
                for col_index in range(1, 14):
                    sheet.cell(row_index, col_index).value = ""
            row_index = index_start
            col_index = 1

            m_rfoods = [f for f in rfoods if f.name == food_name]
            m_cfoods = [f for f in cfoods if f.name == food_name]

            self.unmerge_sheet_cells(sheet)

            sheet.cell(index_start - 2, 1, f"{year}年")

            if len(m_rfoods) > 0:
                for m_row_index in range(
                    index_start, index_start + len(m_rfoods)
                ):
                    food = m_rfoods[m_row_index - index_start]
                    sheet.cell(
                        m_row_index,
                        3,
                        ("上年结转" if month == 1 else "上月结转"),
                    )
                    sheet.cell(row_index, 10, food.count)
                    sheet.cell(row_index, 11, food.unit_price)
                    sheet.cell(row_index, 12, food.count * food.unit_price)
                    row_index += 1
            else:
                sheet.cell(
                    row_index,
                    3,
                    ("上年结转" if month == 1 else "上月结转"),
                )

                row_index += 1

            row_count_form = index_end - index_start - 1

            cdates = []
            for food in m_cfoods:
                if len(food.consumptions) > 0:
                    cdates += [d for d, c in food.consumptions]
                cdates.append(food.xdate)
            cdates = sorted(list(set(cdates)))

            consuming_n = None
            warehousing_n = None

            row_offset = 0

            row_count_foods = 0
            for f in m_cfoods:
                row_count_foods += len([d for d, __ in f.consumptions])
            row_count_foods += len([f.xdate for f in m_cfoods])

            row_count_diff = row_count_foods - row_count_form
            if row_count_diff > 0:
                self.row_inserting_tip(row_index + 1)
                sheet.insert_rows(row_index + 1, row_count_diff + 1)
                for col_index in range(1, 14):
                    for row_index_b in range(
                        row_index + 1, row_index + 1 + row_count_diff + 1
                    ):

                        cell = sheet.cell(row_index_b, col_index)
                        cell.border = self.cell_border0
                        cell.alignment = self.cell_alignment0

                pass

            for cdate in cdates:

                for food in m_cfoods:

                    if food.xdate == cdate and food.xdate.month == month:
                        warehousing_n = warehousing_days.index(cdate) + 1
                        sheet.cell(row_index, 1, cdate.month)
                        sheet.cell(row_index, 2, cdate.day)
                        sheet.cell(row_index, 4, food.count)
                        sheet.cell(row_index, 5, food.unit_price)
                        sheet.cell(row_index, 6, food.count * food.unit_price)
                        sheet.cell(row_index, 9, "")
                        sheet.cell(row_index, 10).number_format = (
                            numbers.FORMAT_NUMBER_00
                        )
                        sheet.cell(row_index, 10, food.count)

                        sheet.cell(row_index, 11, food.unit_price)
                        sheet.cell(row_index, 12, food.count * food.unit_price)
                        sheet.cell(
                            row_index,
                            13,
                            f"R{cdate.month:0>2}{warehousing_n:0>2}",
                        )
                        warehousing_n += 1
                        row_index += 1

                    if cdate in [d for d, __ in food.consumptions]:

                        unit_price = food.unit_price
                        remainder = food.get_remainder(cdate)
                        consuming_n = consuming_days.index(cdate) + 1
                        ccount = [
                            c for d, c in food.consumptions if d == cdate
                        ][0]
                        sheet.cell(row_index, 1, cdate.month)
                        sheet.cell(row_index, 2, cdate.day)
                        sheet.cell(row_index, 6, "")
                        sheet.cell(row_index, 7).number_format = (
                            numbers.FORMAT_NUMBER_00
                        )
                        sheet.cell(row_index, 7, ccount)

                        sheet.cell(row_index, 8, unit_price)
                        sheet.cell(row_index, 9, ccount * unit_price)
                        sheet.cell(row_index, 10, remainder)
                        sheet.cell(
                            row_index,
                            11,
                            food.unit_price,
                        )
                        sheet.cell(row_index, 12, remainder * unit_price)

                        sheet.cell(
                            row_index,
                            13,
                            f"C{cdate.month:0>2}{consuming_n:0>2}",
                        )
                        consuming_n += 1
                        row_index += 1

            self.update_summation(sheet)
            self.update_inventories(sheet)

            self.format(sheet)
            self.update_food_sheet_year(sheet)
            print_info(_("Sheet '%s' was updated.") % sheet.title)

        wb[self.sheet_name].sheet_state = "hidden"

        wb.active = sheet

        bfood_names = list(set([f.name for f in self.bfoods]))
        for name in bfood_names:
            if name in self.bwb.sheetnames:
                sheet = self.get_sheet(name)
                sheet.sheet_properties.tabColor = "0" * 8

        print_info(_("All food sheets have their tab colors reset."))

        for name in food_names:
            sheet = self.get_sheet(name)
            sheet.sheet_properties.tabColor = secrets.token_hex(4)
        food_names_s = sqr_slist(food_names)
        sep_s = get_random_sep_char() * max(
            [
                len(s) + len([c for c in s if is_zh_CN_char(c)])
                for s in food_names_s.split("\n")
            ]
        )

        print_error(sep_s)
        print_warning(_("Food sheets have their tab colors recolor:"))
        print_info(food_names_s)

        print_error(sep_s)
        print_warning(_("Updated food sheets:"))
        print_info(food_names_s)
        print_error(sep_s)


# The end
