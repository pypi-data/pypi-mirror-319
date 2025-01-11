import os
import sys
from openpyxl.styles import *
from openpyxl.formatting.rule import *
from openpyxl.styles.differential import *
from openpyxl.utils.cell import *


from fnschool import *
from fnschool.canteen.spreadsheet.base import *


class Unwarehousing(Base):
    def __init__(self, bill):
        super().__init__(bill)
        self.sheet_name = self.s.unwarehousing_name
        self.entry_row_len0 = 11
        self.form_title0 = "未入库明细表"
        pass

    def get_entry_index(self, form_index):
        form_index0, form_index1 = form_index
        entry_index = [form_index0 + 2, form_index1 - 1]
        return entry_index

    @property
    def form_indexes(self):
        if not self._form_indexes:
            unwsheet = self.sheet
            indexes = []
            row_index = 1
            for row in unwsheet.iter_rows(
                max_row=unwsheet.max_row + 1, max_col=7
            ):
                if row[0].value and self.form_title0 in row[0].value.replace(
                    " ", ""
                ):
                    indexes.append([row_index + 1, 0])

                r1_value = row[1].value
                if r1_value and (
                    "合计" in r1_value.replace(" ", "")
                    or "总合计" in r1_value.replace(" ", "")
                ):
                    indexes[-1][1] = row_index

                row_index += 1

            if len(indexes) > 0:
                self._form_indexes = indexes
            else:
                return None

        return self._form_indexes

    def get_form_title(self, fclass="调味类"):
        title = _("{0} ({1})").format(self.form_title0, fclass)
        return title

    def update(self):
        unwsheet = self.sheet
        form_indexes = self.form_indexes

        ufoods = [f for f in self.bfoods if f.is_abandoned]
        if len(ufoods) < 1:
            print_info(
                _(
                    "There is no abandoned food. "
                    + "Sheet {0} updating skipped."
                ).format(self.sheet.title)
            )
            return None

        food_classes = list(set([f.fclass for f in ufoods]))
        used_form_index0 = None
        used_form_index1 = None
        for fclass in food_classes:
            foods = [f for f in ufoods if f.fclass == fclass]
            foods = sorted(foods, key=lambda f: f.xdate)
            foods_len = len(foods)

            t1 = self.bill.consuming.date_m1
            c_row_count = 0
            used_form_index0 = used_form_index1 + 1 if used_form_index1 else 0
            used_form_index1 = (
                -1 if used_form_index1 == None else used_form_index1
            )
            c_form_indexes = form_indexes[used_form_index0:]
            for i0, i1 in c_form_indexes:
                c_row_count += (i1 - 1) - (i0 + 1)
                used_form_index1 += 1
                if c_row_count >= foods_len:
                    break
            c_form_indexes = form_indexes[
                used_form_index0 : used_form_index1 + 1
            ]

            row_indexes = []
            for form_index in c_form_indexes:
                form_index0, form_index1 = form_index
                unwsheet.cell(form_index0 - 1, 1, self.get_form_title(fclass))
                unwsheet.cell(form_index0, 1, f" 学校名称：{self.purchaser}")
                unwsheet.cell(
                    form_index0,
                    4,
                    f"        "
                    + f"{t1.year} 年 {t1.month} 月 "
                    + f"{t1.day} 日"
                    + f"               ",
                )
                row_index_start = form_index0 + 2
                row_index_end = form_index1 - 1
                row_indexes += list(range(row_index_start, row_index_end + 1))

            for row_index in row_indexes:
                for col_index in range(1, 7 + 1):
                    unwsheet.cell(row_index, col_index, "")

            total_price = 0.0
            cross_forms = False

            for _index, row_index in enumerate(row_indexes):
                food = foods[_index]
                total_price += food.total_price
                unwsheet.cell(row_index, 1, food.xdate.strftime("%Y.%m.%d"))
                unwsheet.cell(row_index, 2, food.name)
                unwsheet.cell(row_index, 3, food.unit_name)

                unwsheet.cell(row_index, 4).number_format = (
                    numbers.FORMAT_NUMBER_00
                )
                unwsheet.cell(row_index, 4, food.count)

                unwsheet.cell(row_index, 5).number_format = (
                    numbers.FORMAT_NUMBER_00
                )
                unwsheet.cell(row_index, 6).number_format = (
                    numbers.FORMAT_NUMBER_00
                )

                unwsheet.cell(row_index, 5, food.unit_price)
                unwsheet.cell(row_index, 6, food.total_price)

                for u_col_index in range(1, 7):
                    cell = unwsheet.cell(row_index, u_col_index)
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

                if (
                    str(unwsheet.cell(row_index + 1, 2).value)
                    .replace(" ", "")
                    .endswith("合计")
                    and foods_len > _index
                ):
                    unwsheet.cell(row_index + 1, 2, "合计")
                    unwsheet.cell(row_index + 1, 6, total_price)
                    cross_forms = True

                if len(foods) - 1 == _index:
                    for row in unwsheet.iter_rows(
                        min_row=row_index,
                        max_row=unwsheet.max_row,
                        min_col=1,
                        max_col=7,
                    ):
                        r1_value = row[1].value
                        if r1_value and str(r1_value).replace(" ", "").endswith(
                            "合计"
                        ):
                            row[1].value = "总合计" if cross_forms else "合计"
                            row[5].value = total_price
                            break
                    break

            print_info(_("Sheet '%s' was updated.") % self.sheet.title)


# The end.
