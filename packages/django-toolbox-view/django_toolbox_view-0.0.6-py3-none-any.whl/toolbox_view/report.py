import datetime
import os

from django.core.files.storage import default_storage

try:
    from openpyxl import Workbook
except ImportError:
    pass


class ReportFile:

    def __init__(self, filename, xl=False, sheet=None, folder='reports/'):
        self.filename = filename
        self.xl = xl
        self.open = True
        self.report_folder = folder
        if xl:
            self.report = Workbook()
            self.sheet = self.report.active
            if sheet:
                self.sheet.title = sheet
        else:
            self.report = self.open_report()

    def open_report(self, write_type='w'):
        report_name, extension = os.path.splitext(self.filename)
        t = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
        return default_storage.open(f'{self.report_folder}{report_name}_{t}{extension}', write_type)

    def get_sheet(self, title):
        if title not in self.report.sheetnames:
            self.sheet = self.report.create_sheet(title)
        else:
            self.sheet = self.report[title]

    def write(self, text):
        if self.xl:
            self.sheet.append(text)
        else:
            self.report.write(text)

    def write_ln(self, text):
        if self.xl:
            self.sheet.append([text])
        else:
            self.report.write(text + '\n')

    def write_csv_line(self, fields):
        if self.xl:
            self.sheet.append(fields)
        else:
            self.report.write(','.join([str(f) for f in fields]) + '\n')

    def close(self):
        self.open = False
        if self.xl:
            f = self.open_report('wb')
            self.report.save(f)
            f.close()
        else:
            self.report.close()

    def __del__(self):
        if self.open:
            self.close()
