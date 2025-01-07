import datetime
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def get_excel_source(source=None):
    if source is None:
        xlsx_files = [f for f in os.listdir(".") if f.endswith(".xlsx")]
        if not xlsx_files:
            raise FileNotFoundError("No .xlsx file found in the current folder.")
        source = xlsx_files[0]
    return source


def parse_excel_data(source):
    try:
        excel_data = pd.ExcelFile(source)
        sheet_name = excel_data.sheet_names[0]
        return excel_data.parse(sheet_name)
    except Exception as e:
        raise ValueError(f"Error parsing Excel file '{source}': {e}")


import os
from datetime import datetime
from openpyxl import Workbook


def create_excel(site_data):
    workbook = Workbook()
    sheet = workbook.active

    standard_font = Font(name="TH Sarabun New", size=16)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    site_is_down_font = Font(bold=True, name="TH Sarabun New", size=16, color="FFFFFF")
    site_is_down_background = PatternFill(start_color="660000", end_color="660000", fill_type="solid")

    fill_even = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    up_status = ["200", "204", "302", "401", "403"]

    headers = ["URL", "Status", "Last Updated"]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, name="TH Sarabun New", size=20)
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, site in enumerate(site_data, start=2):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx in range(1, 4):
            sheet.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, value in enumerate([site.url, site.status, site.updated_at], start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:
                if site.status not in up_status:
                    cell.font = site_is_down_font
                    cell.fill = site_is_down_background
                else:
                    cell.font = standard_font
                    cell.fill = fill_even
            else:
                cell.font = standard_font

            cell.alignment = center_align if col_idx in [2, 3] else left_align
            cell.border = thin_border

    column_widths = [150, 20, 55]
    for col_idx, width in enumerate(column_widths, start=1):
        col_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[col_letter].width = width

    export_folder = ".\\exports"
    os.makedirs(export_folder, exist_ok=True)

    workbook_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_sites_status.xlsx"
    workbook_path = os.path.join(export_folder, workbook_name)
    workbook.save(workbook_path)

    print(f"Excel file created: {workbook_path}")
