import importlib.resources
import os
import sys
import tempfile

import openpyxl
import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_info_cell(worksheet, row, column, text):
    cell = worksheet.cell(row=row, column=column, value=text)

    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.border = thin_border

    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    column_letter = get_column_letter(column)
    max_text_length = max(len(line) for line in text.split("\n"))
    worksheet.column_dimensions[column_letter].width = max_text_length

    return cell


def create_excel_sheet(data_to_export, board_name, output_dir):
    data_export_df = pd.DataFrame(data_to_export)
    try:
        with importlib.resources.open_binary(
            __package__, "csv/trello_template.xlsx"
        ) as template_file:
            workbook = openpyxl.load_workbook(template_file)

    except FileNotFoundError:
        logger.error("Template file not found")
        sys.exit(1)
    except Exception as e:
        logger.error(f"An error occurred while opening the workbook: {str(e)}")
        sys.exit(1)

    worksheet = workbook.active
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    current_list_value = None
    n_row = 1

    for _, card in data_export_df.iterrows():
        list_value = card["List"]

        if list_value != current_list_value:
            current_list_value = list_value
            n_row += 1
            for col in range(1, 7):
                cell = worksheet.cell(row=n_row, column=col)
                cell.fill = grey_fill

        n_row += 1

        for col in [1, 6]:
            cell = worksheet.cell(row=n_row, column=col)
            cell.fill = grey_fill

        worksheet.cell(row=n_row, column=2, value=card["List"])
        worksheet.cell(row=n_row, column=3, value=card["Name"]).alignment = (
            text_alignment
        )
        worksheet.cell(row=n_row, column=4, value=card["Description"]).alignment = (
            text_alignment
        )
        worksheet.cell(row=n_row, column=5, value=card["Labels"]).alignment = (
            text_alignment
        )

    # Insert a closing grey row at the very end
    n_row += 1
    for col in range(1, 7):
        cell = worksheet.cell(row=n_row, column=col)
        cell.fill = grey_fill

    sanitized_board_name = "".join(
        c for c in board_name if c.isalnum() or c in (" ", "_")
    )
    logger.info(f"Sanitized board name: {sanitized_board_name}")

    # Handle saving to local path or temp file
    if output_dir.startswith("s3://"):
        # If it's an S3 URI, save the file to a temporary location
        temp_file = tempfile.NamedTemporaryFile(delete=False, mode="wb")
        temp_filename = temp_file.name
        try:
            workbook.save(temp_file)
            temp_file.close()
            logger.info(f"Excel file created in temporary location: {temp_filename}")
            return temp_filename
        except Exception as e:
            logger.error(f"An error occurred while saving the file: {str(e)}")
            return False
    else:
        # If it's a local path, save it to the specified directory
        sanitized_filename = os.path.join(
            output_dir, f"{sanitized_board_name}_trello_template.xlsx"
        )
        try:
            workbook.save(sanitized_filename)
            logger.info(f"File saved successfully: {sanitized_filename}")
            return sanitized_filename  # Return the local file path
        except Exception as e:
            logger.error(f"An error occurred while saving the file: {str(e)}")
            return False
