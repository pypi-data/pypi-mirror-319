"""Support for Excel parser."""
from typing import  List, Dict
import logging
from datetime import datetime, time,timedelta
import pytz
import dateparser
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl import load_workbook
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta
from pygazpar.enum import NatureReleve, QualificationReleve, StatusReleve,Frequency,PropertyName
from pygazpar.types.RelevesResultType import RelevesResultType
from pygazpar.types.ConsommationType import RelevesType
FIRST_DATA_LINE_NUMBER = 10

Logger = logging.getLogger(__name__)


# ------------------------------------------------------------------------------------------------------------
class ExcelParser:
    '''Excel parser for releve'''
    OUTPUT_DATE_FORMAT = "%Y-%m-%d"
    OUTPUT_DATETIME_FORMAT = "%Y-%m-%dT%H:%M:%S%z"

    INPUT_DATE_FORMAT = "%d/%m/%Y"
    # ------------------------------------------------------
    @staticmethod
    def parse(data_filename: str, data_reading_frequency: Frequency) -> List[RelevesResultType]:
        '''Parse excel file'''
        parse_by_frequency = {
            Frequency.HOURLY: ExcelParser.__parse_hourly,
            Frequency.DAILY: ExcelParser.__parse_daily,
            Frequency.WEEKLY: ExcelParser.__parse_weekly,
            Frequency.MONTHLY: ExcelParser.__parse_monthly
        }

        Logger.debug(f"Loading Excel data file '{data_filename}'...")

        workbook = load_workbook(filename=data_filename)

        worksheet = workbook.active

        res = parse_by_frequency[data_reading_frequency](worksheet)  # type: ignore

        workbook.close()

        return res

    # ------------------------------------------------------
    @staticmethod
    def __fill_row(row: Dict, property_name: str, cell: Cell, is_number: bool):
        '''fill row dictionnary with value from excel'''
        if cell.value is not None:
            if is_number:
                if isinstance(cell.value,str):
                    if len(cell.value.strip()) > 0:
                        row[property_name] = float(cell.value.replace(',', '.'))
                else:
                    row[property_name] = cell.value
            else:
                row[property_name] = cell.value.strip() if isinstance(cell.value,str) else cell.value
        else:
            row[property_name] = None

    # ------------------------------------------------------
    @staticmethod
    def __parse_hourly(worksheet: Worksheet) -> List[RelevesResultType]:
        '''Parse hourly data'''
        return []

    # ------------------------------------------------------
    @staticmethod
    def __parse_daily(worksheet: Worksheet) -> List[RelevesResultType]:
        '''parse daily data'''
        res = []       
        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()

        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                date_journee = datetime.strptime(worksheet.cell(column=2, row=rownum).value, ExcelParser.INPUT_DATE_FORMAT).date()
                info=pytz.timezone('Europe/Paris')
                MyTime = time(6, 0, 0)  #hr/min/sec
                datetime_debut = datetime.combine(date_journee, MyTime)
                datetime_debut_localize=info.localize(datetime_debut)
                row[PropertyName.JOURNEE_GAZIERE.value] = date_journee.strftime(ExcelParser.OUTPUT_DATE_FORMAT)
                row[PropertyName.DATE_DEBUT.value]= datetime_debut_localize.isoformat()
                row[PropertyName.DATE_FIN.value]= (datetime_debut_localize+timedelta(days=1)).isoformat()

                ExcelParser.__fill_row(row, PropertyName.START_INDEX.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.END_INDEX.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.VOLUME.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.ENERGY.value, worksheet.cell(column=6, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.CONVERTER_FACTOR.value, worksheet.cell(column=7, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=8, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.QUALIFICATION.value, worksheet.cell(column=9, row=rownum), False)  # type: ignore
                row[PropertyName.PCS.value]=None
                row[PropertyName.VOLUME_CONVERTI.value]=round(row[PropertyName.VOLUME.value])
                row[PropertyName.PTA.value]=None
                row[PropertyName.NATURE.value]=NatureReleve.INFORMATIVES.value
                row[PropertyName.STATUS.value]=StatusReleve.PROVISOIRE.value
                row[PropertyName.FREQUENCE_RELEVE.value]=None
                releve = RelevesType(**row)
                releve_result = RelevesResultType(worksheet.cell(column=2, row=rownum).value,data_timestamp,releve)

                res.append(releve_result)

        Logger.debug(f"Daily data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res

    # ------------------------------------------------------
    @staticmethod
    def __parse_weekly(worksheet: Worksheet) -> List[RelevesResultType]:
        '''parse weekly data'''
        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()
        info=pytz.timezone('Europe/Paris')
        my_time = time(6, 0, 0)  #hr/min/sec
        min_row_num = FIRST_DATA_LINE_NUMBER
        max_row_num = len(worksheet['B'])
        for rownum in range(min_row_num, max_row_num + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                dateField=worksheet.cell(column=2, row=rownum).value
                dateStart=dateField.split('au')[0]
                dateEnd=dateField.split('au')[1]
                dateStartDT=parse(dateStart, fuzzy_with_tokens=True)
                dateEndDT=parse(dateEnd, fuzzy_with_tokens=True)
                dateStartDT = datetime.combine(dateStartDT[0], my_time)
                dateEndDT = datetime.combine(dateEndDT[0], my_time)
                dateStartDT=info.localize(dateStartDT)
                dateEndDT=info.localize(dateEndDT)

                row[PropertyName.DATE_DEBUT.value]= dateStartDT.isoformat()
                row[PropertyName.DATE_FIN.value]= (dateEndDT+timedelta(days=1)).isoformat()
                row[PropertyName.JOURNEE_GAZIERE.value] =None
                ExcelParser.__fill_row(row, PropertyName.VOLUME.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.ENERGY.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore

                row[PropertyName.START_INDEX.value]=None
                row[PropertyName.END_INDEX.value]=None
                row[PropertyName.CONVERTER_FACTOR.value]=None
                row[PropertyName.QUALIFICATION.value]=QualificationReleve.ESTIME.value

                row[PropertyName.PCS.value]=None
                row[PropertyName.VOLUME_CONVERTI.value]=round(row[PropertyName.VOLUME.value])
                row[PropertyName.PTA.value]=None
                row[PropertyName.NATURE.value]=NatureReleve.INFORMATIVES.value
                row[PropertyName.STATUS.value]=StatusReleve.PROVISOIRE.value
                row[PropertyName.FREQUENCE_RELEVE.value]=None
                releve = RelevesType(**row)
                releve_result = RelevesResultType(worksheet.cell(column=2, row=rownum).value,data_timestamp,releve)
                res.append(releve_result)

        Logger.debug(f"Weekly data read successfully between row #{min_row_num} and row #{max_row_num}")

        return res

    # ------------------------------------------------------
    @staticmethod
    def __parse_monthly(worksheet: Worksheet) -> List[RelevesResultType]:
        '''parse Monthly data'''
        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()
        info=pytz.timezone('Europe/Paris')
       
        MyTime = time(6, 0, 0)  #hr/min/sec
        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                dateField=worksheet.cell(column=2, row=rownum).value
                dateStartDT=dateparser.parse(dateField, locales=['fr'])
                dateStartDT=dateStartDT.replace(day=1)
                dateStartDT = datetime.combine(dateStartDT, MyTime)
                dateStartDT=info.localize(dateStartDT)
                row[PropertyName.DATE_DEBUT.value]= dateStartDT.isoformat()
                row[PropertyName.DATE_FIN.value]= (dateStartDT+relativedelta(months=1)).isoformat()
                row[PropertyName.JOURNEE_GAZIERE.value] =None

                ExcelParser.__fill_row(row, PropertyName.VOLUME.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.ENERGY.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fill_row(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore

                row[PropertyName.START_INDEX.value]=None
                row[PropertyName.END_INDEX.value]=None
                row[PropertyName.CONVERTER_FACTOR.value]=None
                row[PropertyName.QUALIFICATION.value]=QualificationReleve.ESTIME.value
                row[PropertyName.PCS.value]=None
                row[PropertyName.VOLUME_CONVERTI.value]=round(row[PropertyName.VOLUME.value])
                row[PropertyName.PTA.value]=None
                row[PropertyName.NATURE.value]=NatureReleve.INFORMATIVES.value
                row[PropertyName.STATUS.value]=StatusReleve.PROVISOIRE.value
                row[PropertyName.FREQUENCE_RELEVE.value]=None
                
                releve = RelevesType(**row)
                releve_result = RelevesResultType(worksheet.cell(column=2, row=rownum).value,data_timestamp,releve)
                res.append(releve_result)

        Logger.debug(f"Monthly data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res
